import torch
import torch.nn as nn
import numpy as np
import pandas as pd
from sklearn.metrics.pairwise import euclidean_distances
from sklearn.datasets import fetch_openml
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.ensemble import AdaBoostClassifier, BaggingClassifier, RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.naive_bayes import GaussianNB
from sklearn.svm import SVC
from sklearn.metrics import accuracy_score, f1_score
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Define a simple fully connected network for feature extraction
class SimpleFeatureExtractor(nn.Module):
    def __init__(self, input_dim, hidden_dim=16, output_dim=8):
        super(SimpleFeatureExtractor, self).__init__()
        self.fc1 = nn.Linear(input_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, output_dim)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        return x

# Core-Set Selection Function using K-Center Greedy
def k_center_greedy(X, num_samples):
    selected_indices = [np.random.randint(len(X))]
    min_distances = euclidean_distances(X, X[selected_indices]).min(axis=1)

    for _ in tqdm(range(num_samples - 1), leave=False):
        farthest_idx = np.argmax(min_distances)
        selected_indices.append(farthest_idx)
        distances_to_new = euclidean_distances(X, X[[farthest_idx]]).flatten()
        min_distances = np.minimum(min_distances, distances_to_new)

    return selected_indices

# Run experiment on the Balance dataset
def run_balance_experiment():
    # Load the Balance Scale dataset
    dataset = fetch_openml(data_id=11)  # Balance Scale dataset ID
    X = dataset.data
    y = dataset.target

    # Encode target labels to numeric values
    label_encoder = LabelEncoder()
    y = label_encoder.fit_transform(y)

    # Standardize features
    scaler = StandardScaler()
    X = scaler.fit_transform(X)

    # Convert data to torch tensors for feature extraction
    X_tensor = torch.tensor(X, dtype=torch.float32)

    # Initialize the feature extractor model
    input_dim = X.shape[1]
    feature_extractor = SimpleFeatureExtractor(input_dim=input_dim)

    # Define classifiers
    classifiers = {
        "AdaBoost": AdaBoostClassifier(algorithm="SAMME"),
        "Bagging": BaggingClassifier(),
        "Decision Tree": DecisionTreeClassifier(),
        "KNN": KNeighborsClassifier(),
        "Logistic Regression": LogisticRegression(),
        "Naive Bayes": GaussianNB(),
        "Random Forest": RandomForestClassifier(),
        "SVM": SVC()
    }

    # Parameters for the experiment
    num_runs = 30
    num_samples = int(0.1 * len(X))  # 10% of the dataset
    results = {classifier: {"Accuracy": [], "F1 Score": []} for classifier in classifiers.keys()}
    times = []
    core_set_indices_all_runs = []
    best_combination_scores = []

    # Run the experiment 30 times
    for run in range(num_runs):
        # Extract features
        with torch.no_grad():
            feature_extractor.eval()
            X_features = feature_extractor(X_tensor).numpy()  # Extract features

        # core-set selection
        core_set_indices = k_center_greedy(X_features, num_samples)

        # Record core-set indices
        core_set_indices_all_runs.append(core_set_indices)

        # Get core-set data and labels
        core_set_data = X[core_set_indices]
        core_set_labels = y[core_set_indices]

        # Remaining data and labels
        remaining_indices = list(set(range(len(X))) - set(core_set_indices))
        remaining_data = X[remaining_indices]
        remaining_labels = y[remaining_indices]

        # Train and evaluate each classifier, recording combined scores
        combined_scores = []
        for clf_name, clf in classifiers.items():
            clf.fit(core_set_data, core_set_labels)
            y_pred_remaining = clf.predict(remaining_data)

            # Calculate metrics
            accuracy = accuracy_score(remaining_labels, y_pred_remaining)
            f1 = f1_score(remaining_labels, y_pred_remaining, average='macro')
            results[clf_name]["Accuracy"].append(accuracy)
            results[clf_name]["F1 Score"].append(f1)
            
            combined_score = (accuracy + f1) / 2  # Combine accuracy and F1 for this run and classifier
            combined_scores.append(combined_score)
        
        # Store the best combined score for each run
        best_combination_scores.append(max(combined_scores))

    # Identify the run with the best combined accuracy and F1 score
    best_run_index = np.argmax(best_combination_scores)
    best_core_set_indices = core_set_indices_all_runs[best_run_index]

    # Prepare the results DataFrame
    summary_data = []

    for clf_name, metrics in results.items():
        accuracy_best = np.max(metrics["Accuracy"])
        accuracy_worst = np.min(metrics["Accuracy"])
        accuracy_avg = np.mean(metrics["Accuracy"])
        accuracy_std = np.std(metrics["Accuracy"])

        f1_best = np.max(metrics["F1 Score"])
        f1_worst = np.min(metrics["F1 Score"])
        f1_avg = np.mean(metrics["F1 Score"])
        f1_std = np.std(metrics["F1 Score"])

        summary_data.append({
            "Classifier": clf_name,
            "Accuracy Best": accuracy_best,
            "Accuracy Worst": accuracy_worst,
            "Accuracy Average": accuracy_avg,
            "Accuracy Std Dev": accuracy_std,
            "F1 Score Best": f1_best,
            "F1 Score Worst": f1_worst,
            "F1 Score Average": f1_avg,
            "F1 Score Std Dev": f1_std,
        })

    summary_df = pd.DataFrame(summary_data)

    # Create DataFrame for core-set indices for each run (transposed to align columns correctly)
    core_set_indices_df = pd.DataFrame(core_set_indices_all_runs).transpose()
    core_set_indices_df.columns = [f"Run {i+1}" for i in range(num_runs)]

    # Save results to an Excel file named after the dataset
    filename = "Balance_experiment_results.xlsx"
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        core_set_indices_df.to_excel(writer, sheet_name="Core Set Indices", index=False)

        # Highlight the best core-set indices in the Excel file
        workbook = writer.book
        sheet = workbook["Core Set Indices"]

        # Apply highlight to the best run's core-set indices
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in range(2, len(best_core_set_indices) + 2):  # +2 to account for header
            cell = sheet.cell(row=row, column=best_run_index + 2)  # +2 for header and 1-based indexing
            cell.fill = highlight_fill

    print(f"Experiment results for Balance dataset saved to {filename}, with the best core-set indices highlighted.")

# Run the experiment on the Balance dataset
run_balance_experiment()
