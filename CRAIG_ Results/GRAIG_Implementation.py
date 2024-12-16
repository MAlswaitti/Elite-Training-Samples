import os
import random
import numpy as np
import heapq
import pandas as pd
from sklearn.ensemble import AdaBoostClassifier, BaggingClassifier, RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.naive_bayes import GaussianNB
from sklearn.svm import SVC
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import accuracy_score, f1_score
from scipy.spatial.distance import cdist
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import sklearn.utils

# Simplified Facility Location class
class FacilityLocation:
    def __init__(self, S, V):
        self.S = S
        self.V = V
        self.curVal = 0
        self.curMax = np.zeros(len(S))
    
    def inc(self, X, Y):
        return np.maximum(self.curMax, self.S[X])

    def gain(self, X, Y):
        return np.sum(self.inc(X, Y)) - self.curVal

    def add(self, X):
        self.curMax = self.inc(X, self.curMax)
        self.curVal = np.sum(self.curMax)

# Lazy Greedy algorithm
def lazy_greedy_heap(F, V, B):
    heap = [(-F.gain(i, []), i) for i in V]
    heapq.heapify(heap)
    solution = []
    for _ in range(B):
        _, x = heapq.heappop(heap)
        solution.append(x)
        F.add(x)
        heap = [(-F.gain(i, solution), i) for _, i in heap]
        heapq.heapify(heap)
    return solution, F.curVal

# Coreset Selection using CRAIG
def craig_coreset(X, fraction=0.1):
    N = len(X)
    B = int(fraction * N)
    V = list(range(N))
    S = 1 - cdist(X, X, metric='cosine')  # Similarity matrix
    F = FacilityLocation(S, V)
    order, _ = lazy_greedy_heap(F, V, B)
    return order

# Define classifiers
classifiers = {
    "AdaBoost": AdaBoostClassifier(algorithm="SAMME"),
    "Bagging": BaggingClassifier(),
    "Decision Tree": DecisionTreeClassifier(),
    "KNN": KNeighborsClassifier(),
    "Logistic Regression": LogisticRegression(),
    "Naive Bayes": GaussianNB(),
    "Random Forest": RandomForestClassifier(),
    "SVM": SVC()
}

# Perform the experiment
def run_experiment(csv_path, runs=30, fraction=0.1):
    dataset_name = os.path.splitext(os.path.basename(csv_path))[0]
    data = pd.read_csv(csv_path)
    X = data.iloc[:, :-1].values
    y = data.iloc[:, -1].values

    scaler = StandardScaler()

    results = {clf_name: {"Accuracy": [], "F1 Score": []} for clf_name in classifiers.keys()}
    times = []
    core_set_indices_all_runs = []

    for run in range(runs):
        # Shuffle dataset for randomness
        X, y = sklearn.utils.shuffle(X, y, random_state=None)
        X = scaler.fit_transform(X)

        coreset_indices = craig_coreset(X, fraction=fraction)

        times.append(end_time - start_time)
        core_set_indices_all_runs.append(coreset_indices)

        X_coreset = X[coreset_indices]
        y_coreset = y[coreset_indices]
        test_indices = list(set(range(len(X))) - set(coreset_indices))
        X_test = X[test_indices]
        y_test = y[test_indices]

        for clf_name, clf in classifiers.items():
            clf.fit(X_coreset, y_coreset)
            y_pred = clf.predict(X_test)

            accuracy = accuracy_score(y_test, y_pred)
            f1 = f1_score(y_test, y_pred, average="macro")
            results[clf_name]["Accuracy"].append(accuracy)
            results[clf_name]["F1 Score"].append(f1)

    summary_data = []
    for clf_name, metrics in results.items():
        accuracy_best = np.max(metrics["Accuracy"])
        accuracy_worst = np.min(metrics["Accuracy"])
        accuracy_avg = np.mean(metrics["Accuracy"])
        accuracy_std = np.std(metrics["Accuracy"])

        f1_best = np.max(metrics["F1 Score"])
        f1_worst = np.min(metrics["F1 Score"])
        f1_avg = np.mean(metrics["F1 Score"])
        f1_std = np.std(metrics["F1 Score"])

        summary_data.append({
            "Classifier": clf_name,
            "Accuracy Best": accuracy_best,
            "Accuracy Worst": accuracy_worst,
            "Accuracy Average": accuracy_avg,
            "Accuracy Std Dev": accuracy_std,
            "F1 Score Best": f1_best,
            "F1 Score Worst": f1_worst,
            "F1 Score Average": f1_avg,
            "F1 Score Std Dev": f1_std,
        })

    summary_df = pd.DataFrame(summary_data)
    core_set_indices_df = pd.DataFrame(core_set_indices_all_runs).transpose()
    core_set_indices_df.columns = [f"Run {i+1}" for i in range(runs)]

    filename = f"{dataset_name}_experiment_results.xlsx"
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        core_set_indices_df.to_excel(writer, sheet_name="Core Set Indices", index=False)

        workbook = writer.book
        sheet = workbook["Core Set Indices"]

        best_run_index = np.argmax(summary_df["Accuracy Best"].values)
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in range(2, len(core_set_indices_all_runs[best_run_index]) + 2):
            cell = sheet.cell(row=row, column=best_run_index + 2)
            cell.fill = highlight_fill

    print(f"Results saved to {filename}.")

# Run the experiment for the Iris dataset (example)
run_experiment('Datasets/iris.csv')
